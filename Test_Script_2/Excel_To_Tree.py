import sys
import json
import pandas as pd
from openpyxl import load_workbook
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QTreeWidget, QTreeWidgetItem, QFileDialog, 
    QVBoxLayout, QWidget, QHBoxLayout, QTextEdit, QLabel, QSplitter
)
from PySide6.QtCore import Qt

class TreeApp(QMainWindow):
    def __init__(self, excel_path):
        super().__init__()
        self.setWindowTitle("Register Tree Viewer")
        self.setGeometry(200, 200, 1200, 800)

        # 메인 위젯과 레이아웃 설정
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        main_layout = QVBoxLayout(main_widget)

        # 수평 스플리터로 트리와 디스크립션 영역 분할
        splitter = QSplitter(Qt.Horizontal)
        main_layout.addWidget(splitter)

        # 트리 위젯
        self.tree = QTreeWidget()
        self.tree.setHeaderLabels(["Register/Field"])
        self.tree.itemClicked.connect(self.on_item_clicked)
        splitter.addWidget(self.tree)

        # 디스크립션 영역
        desc_widget = QWidget()
        desc_layout = QVBoxLayout(desc_widget)
        desc_layout.addWidget(QLabel("Description:"))
        self.desc_text = QTextEdit()
        self.desc_text.setReadOnly(True)
        self.desc_text.setMaximumHeight(200)
        desc_layout.addWidget(self.desc_text)
        splitter.addWidget(desc_widget)

        # 스플리터 비율 설정 (트리:디스크립션 = 3:1)
        splitter.setSizes([900, 300])

        # 데이터 저장용
        self.data = None

        # Load Excel and Build Tree (Excel 파일을 우선적으로 로드)
        try:
            self.data = self.load_excel(excel_path)
            print(f"✅ Excel 파일 로드: {excel_path}")
        except Exception as e:
            print(f"Excel 파일 로드 실패: {e}")
            # 기존 JSON 파일 로드 시도
            json_path = r"c:\Project\Work\Python\FT2232H_SPI_2nd\register_tree.json"
            try:
                with open(json_path, "r", encoding="utf-8") as f:
                    self.data = json.load(f)
                print(f"✅ JSON 파일 로드: {json_path}")
            except Exception as json_e:
                print(f"JSON 파일 로드 실패: {json_e}")
                return

        if self.data:
            self.build_tree(self.data)

            # Save to JSON
            json_path = excel_path.replace(".xlsx", "_tree.json")
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(self.data, f, indent=4, ensure_ascii=False)
            print(f"✅ JSON 파일 저장 완료: {json_path}")
        else:
            print("❌ 데이터를 로드할 수 없습니다.")

    def on_item_clicked(self, item, column):
        """트리 아이템 클릭 시 Description 표시"""
        item_text = item.text(0)
        description = ""
        
        # 레지스터 노드인지 확인 (주소로 시작하는지)
        for reg_addr, reg_data in self.data["Device"].items():
            if item_text.startswith(reg_addr):
                description = f"Register: {reg_data['Description']}\nAddress: {reg_addr}"
                break
        else:
            # 필드 노드 - 부모에서 레지스터 찾기
            parent = item.parent()
            if parent:
                parent_text = parent.text(0)
                for reg_addr, reg_data in self.data["Device"].items():
                    if parent_text.startswith(reg_addr):
                        # 필드명 추출 (= 앞부분)
                        field_name = item_text.split(' = ')[0]
                        if field_name in reg_data:
                            field_data = reg_data[field_name]
                            description = f"Field: {field_name}\n"
                            description += f"Bits: {field_data['upper']}:{field_data['lower']}\n"
                            description += f"Value: {field_data['value']}\n"
                            description += f"Description: {field_data.get('Meaning', 'N/A')}"
                        break
        
        self.desc_text.setText(description)

    def load_excel(self, path):
        """Excel 파일을 읽어서 병합된 셀 정보와 함께 JSON 구조로 변환"""
        # pandas로 데이터 읽기
        df = pd.read_excel(path, header=None)
        
        # openpyxl로 병합된 셀 정보 읽기
        wb = load_workbook(path)
        ws = wb.active
        merged_ranges = ws.merged_cells.ranges
        
        print(f"📊 Excel 파일 크기: {df.shape[0]}행 x {df.shape[1]}열")
        print(f"🔗 병합된 셀 범위: {len(merged_ranges)}개")
        
        data = {"Device": {}}
        
        # 필드별 의미 정보를 저장할 딕셔너리 (Name -> Meaning 매핑)
        field_meanings = {}
        
        # Excel 파일 전체에서 모든 Meaning 열 찾기
        meaning_positions = []
        for row_idx in range(len(df)):
            for col_idx in range(len(df.columns)):
                cell_value = df.iat[row_idx, col_idx]
                if pd.notna(cell_value) and str(cell_value).strip() == "Meaning":
                    meaning_positions.append((row_idx, col_idx))
        
        print(f"🔍 총 {len(meaning_positions)}개의 Meaning 테이블 발견")
        
        # 각 Meaning 테이블에서 정보 수집
        for table_idx, (meaning_row, meaning_col) in enumerate(meaning_positions):
            print(f"� Meaning 테이블 #{table_idx + 1} 처리 중 (Row {meaning_row}, Col {meaning_col})")
            
            # Name 열은 보통 Meaning 열보다 4칸 앞에 위치
            name_col = meaning_col - 4
            if name_col < 0:
                continue
                
            # 해당 테이블의 데이터 행들 처리
            for data_row_idx in range(meaning_row + 1, len(df)):
                if data_row_idx >= len(df):
                    break
                    
                data_row = df.iloc[data_row_idx]
                
                # Name과 Meaning 값 가져오기
                name_val = data_row.iloc[name_col] if name_col < len(data_row) else None
                meaning_val = data_row.iloc[meaning_col] if meaning_col < len(data_row) else None
                
                # 유효한 데이터인지 확인
                if pd.notna(name_val) and pd.notna(meaning_val):
                    name_str = str(name_val).strip()
                    meaning_str = str(meaning_val).strip()
                    
                    # "Name" 헤더가 다시 나오면 다음 테이블 시작이므로 중단
                    if name_str == "Name":
                        break
                        
                    if name_str and meaning_str and meaning_str != "nan":
                        field_meanings[name_str] = meaning_str
                        print(f"   📝 필드 의미: {name_str} = {meaning_str}")
                
                # 빈 행이 연속으로 나오면 테이블 끝
                elif pd.isna(name_val) and pd.isna(meaning_val):
                    # 다음 몇 행도 확인해서 정말 끝인지 체크
                    empty_count = 0
                    for check_row in range(data_row_idx, min(data_row_idx + 3, len(df))):
                        check_data = df.iloc[check_row]
                        if pd.isna(check_data.iloc[name_col]) and pd.isna(check_data.iloc[meaning_col]):
                            empty_count += 1
                    if empty_count >= 2:  # 2행 이상 비어있으면 테이블 끝
                        break
        
        # 병합된 셀 정보를 딕셔너리로 변환 (더 빠른 검색을 위해)
        merged_info = {}
        for merged_range in merged_ranges:
            min_row, min_col = merged_range.min_row - 1, merged_range.min_col - 1  # 0-based 인덱스
            max_row, max_col = merged_range.max_row - 1, merged_range.max_col - 1
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    merged_info[(r, c)] = {
                        'min_row': min_row, 'max_row': max_row,
                        'min_col': min_col, 'max_col': max_col,
                        'is_master': (r == min_row and c == min_col)
                    }
        
        # "Addr" 문자열이 포함된 셀 찾기
        for row_idx, row in df.iterrows():
            for col_idx, cell in enumerate(row):
                if isinstance(cell, str) and "Addr" in cell:
                    print(f"\n🎯 레지스터 발견: Row {row_idx}, Col {col_idx}")
                    
                    addr_col = col_idx
                    reg_col = col_idx + 1
                    desc_col = col_idx + 2
                    
                    # Register 정보 읽기
                    addr_value = str(df.iat[row_idx, addr_col + 1]).strip()  # Addr 옆의 값
                    desc_value = str(df.iat[row_idx, desc_col]).strip()
                    
                    if not addr_value or addr_value == "nan":
                        continue
                    
                    print(f"   📍 주소: {addr_value}, 설명: {desc_value}")
                    
                    reg_dict = {
                        "Description": desc_value
                    }
                    
                    # 비트 행 찾기 (Addr 다음 행에서 "Bit" 찾기)
                    bit_row = None
                    name_row = None
                    default_row = None
                    
                    for search_row in range(row_idx + 1, min(row_idx + 10, len(df))):
                        first_cell = str(df.iat[search_row, col_idx]).strip()
                        if first_cell == "Bit":
                            bit_row = search_row
                        elif first_cell == "Name":
                            name_row = search_row
                        elif first_cell == "Default":
                            default_row = search_row
                            break
                    
                    if bit_row is None or name_row is None or default_row is None:
                        print(f"   ⚠️ 비트 정보를 찾을 수 없음")
                        continue
                    
                    print(f"   📋 Bit행: {bit_row}, Name행: {name_row}, Default행: {default_row}")
                    
                    # 비트 필드 정보 수집
                    bit_fields = {}
                    
                    # 16비트 (0-15) 순회
                    for bit_col in range(addr_col + 1, min(addr_col + 17, len(df.columns))):
                        bit_num_cell = df.iat[bit_row, bit_col]
                        name_cell = df.iat[name_row, bit_col]
                        default_cell = df.iat[default_row, bit_col]
                        
                        # 비트 번호 확인
                        if pd.isna(bit_num_cell):
                            continue
                            
                        try:
                            bit_num = int(bit_num_cell)
                        except:
                            continue
                        
                        # Name 셀 처리 (병합된 셀 고려)
                        field_name = ""
                        if not pd.isna(name_cell):
                            field_name = str(name_cell).strip()
                        
                        # 병합된 셀인지 확인
                        merge_info = merged_info.get((name_row, bit_col))
                        if merge_info:
                            # 병합된 셀의 시작점에서 이름 가져오기
                            master_name = df.iat[merge_info['min_row'], merge_info['min_col']]
                            if not pd.isna(master_name):
                                field_name = str(master_name).strip()
                            
                            # 병합 범위 계산 (비트 번호 기준)
                            upper_bit = 15 - (merge_info['min_col'] - addr_col - 1)
                            lower_bit = 15 - (merge_info['max_col'] - addr_col - 1)
                            
                            # upper가 lower보다 작으면 바꿔줌
                            if upper_bit < lower_bit:
                                upper_bit, lower_bit = lower_bit, upper_bit
                                
                        else:
                            # 단일 비트
                            upper_bit = lower_bit = bit_num
                        
                        # Default 값 처리
                        default_val = 0
                        if not pd.isna(default_cell):
                            try:
                                default_val = int(default_cell)
                            except:
                                default_val = 0
                        
                        # 필드명이 있는 경우만 추가
                        if field_name and field_name not in ["nan", ""]:
                            # 필드명 정리
                            clean_name = field_name.replace("<", "").replace(">", "").replace(":", "_").replace(" ", "_")
                            clean_name = clean_name.replace("'", "").replace("b", "")  # 1'b0 -> 10
                            
                            if clean_name not in bit_fields:
                                # 의미 정보 가져오기 (field_meanings 딕셔너리에서)
                                field_meaning = field_meanings.get(field_name, f"{field_name} bits {upper_bit}:{lower_bit}" if upper_bit != lower_bit else f"{field_name} bit {upper_bit}")
                                
                                bit_fields[clean_name] = {
                                    "upper": upper_bit,
                                    "lower": lower_bit,
                                    "value": default_val,
                                    "Meaning": field_meaning
                                }
                                print(f"     🔹 필드: {clean_name} = bit {upper_bit}:{lower_bit}, 기본값: {default_val}, 의미: {field_meaning}")
                    
                    # 레지스터 딕셔너리에 비트 필드들 추가
                    reg_dict.update(bit_fields)
                    data["Device"][addr_value] = reg_dict
                    print(f"   ✅ 레지스터 {addr_value} 처리 완료 ({len(bit_fields)}개 필드)")
        
        return data

    def build_tree(self, data):
        device_node = QTreeWidgetItem(["RSC201_DR_v0_250418 (002)_UM232H"])
        self.tree.addTopLevelItem(device_node)

        if "Device" not in data:
            return

        for reg_addr, reg_data in data["Device"].items():
            # 레지스터 노드: 주소와 설명 표시
            reg_display = f"[{reg_addr}] {reg_data.get('Description', 'No Description')}"
            reg_node = QTreeWidgetItem([reg_display])
            device_node.addChild(reg_node)

            # 비트 필드들을 개별적으로 표시
            for field_name, field_data in reg_data.items():
                if field_name != "Description" and isinstance(field_data, dict):
                    value = field_data.get("value", 0)
                    # 필드명과 값을 단순하게 표시
                    field_display = f"{field_name} = {value}"
                    field_node = QTreeWidgetItem([field_display])
                    reg_node.addChild(field_node)

        self.tree.expandAll()


if __name__ == "__main__":
    app = QApplication(sys.argv)

    # 특정 엑셀 파일 경로 설정
    file_path = r"c:\Project\Work\Python\FT2232H_SPI_2nd\RSC201_DR_v0_250418 (002)_UM232H.xlsx"
    
    # 파일 존재 확인
    import os
    if os.path.exists(file_path):
        window = TreeApp(file_path)
        window.show()
        sys.exit(app.exec())
    else:
        print(f"❌ 파일을 찾을 수 없습니다: {file_path}")
        # 파일이 없으면 파일 선택 다이얼로그 열기
        file_path, _ = QFileDialog.getOpenFileName(None, "Select Excel File", "", "Excel Files (*.xlsx)")
        if file_path:
            window = TreeApp(file_path)
            window.show()
            sys.exit(app.exec())
        else:
            print("❌ 파일이 선택되지 않았습니다.")