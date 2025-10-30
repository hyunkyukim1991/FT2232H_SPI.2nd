import sys
import json
import pandas as pd
from openpyxl import load_workbook
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QFileDialog, QMessageBox, 
    QPushButton, QHBoxLayout, QVBoxLayout, QGridLayout, QLabel, QTreeWidgetItem, QSpinBox
)
from PySide6.QtCore import Qt, QFile, QIODevice
from PySide6.QtUiTools import QUiLoader

# Custom UInt32 SpinBox 임포트
from uint32_spinbox import UInt32SpinBox

# FT2232H SPI 통신을 위한 import (pyftdi 라이브러리 필요)
try:
    from pyftdi.spi import SpiController
    PYFTDI_AVAILABLE = True
except ImportError:
    PYFTDI_AVAILABLE = False
    print("⚠️ pyftdi 라이브러리가 설치되지 않았습니다. SPI 통신 기능이 제한됩니다.")
    print("설치: pip install pyftdi")

class RegisterTreeViewerController(QMainWindow):
    def __init__(self, excel_path=None):
        super().__init__()
        
        # SPI 컨트롤러 초기화
        self.spi_controller = None
        self.spi = None
        
        # 현재 선택된 레지스터 정보
        self.current_register = None
        self.current_register_data = None
        
        # 현재 선택된 필드 정보
        self.current_field = None
        self.current_field_data = None
        
        # 전역 레지스터 상태 관리
        self.reg_addr = None  # 현재 선택된 레지스터 주소 (문자열, 예: "02", "2B")
        self.reg_data = 0     # 현재 레지스터의 데이터 값 (정수)
        
        # UI 업데이트 동기화 플래그
        self._updating_ui = False
        
        # 비트 버튼들을 저장할 리스트
        self.bit_buttons = []
        
        # 레지스터 데이터
        self.data = None
        
        # UI 로드
        self.load_ui()
        
        # 시그널 연결
        self.connect_signals()
        
        # 32개 비트 버튼 생성
        self.create_bit_buttons()
        
        # 초기 UI 상태 설정
        self.setup_initial_ui_state()
        
        # Excel 파일이 지정되면 로드
        if excel_path:
            self.load_excel_file(excel_path)
            
        # 모든 비트 버튼을 0으로 초기화
        self.reset_all_bits_to_zero()

    def load_ui(self):
        """UI 파일을 로드합니다."""
        ui_file = QFile("register_controller.ui")
        if not ui_file.open(QIODevice.ReadOnly):
            print("UI 파일을 열 수 없습니다.")
            return
            
        loader = QUiLoader()
        self.ui = loader.load(ui_file)
        ui_file.close()
        
        if not self.ui:
            print("UI 로드 실패")
            return
            
        # UI의 모든 위젯들을 현재 MainWindow에 복사
        self.setCentralWidget(self.ui.centralwidget)
        self.setMenuBar(self.ui.menubar)
        self.setStatusBar(self.ui.statusbar)
        
        # 윈도우 속성 설정
        self.setWindowTitle("Register Tree Viewer with 32-bit Controller")
        self.setGeometry(100, 100, 1400, 800)
        
        # QSpinBox 32비트 설정
        self.setup_spinbox()

    def setup_spinbox(self):
        """QSpinBox를 32비트 처리용으로 설정 (간단한 방법)"""
        if hasattr(self.ui, 'hex_value_spinbox'):
            spinbox = self.ui.hex_value_spinbox
            # signed int 범위 설정
            spinbox.setRange(-2147483648, 2147483647)
            spinbox.setDisplayIntegerBase(16)  # 16진수 표시
            spinbox.setPrefix("0x")  # 0x 접두사
            spinbox.setValue(0)  # 초기값
            print("✅ QSpinBox 설정 완료 (32비트 처리 로직 포함)")

    def setup_initial_ui_state(self):
        """초기 UI 상태를 설정합니다."""
        try:
            # 연결 관련 버튼 - Connect는 활성화, Disconnect는 비활성화
            if hasattr(self.ui, 'connect_btn'):
                self.ui.connect_btn.setEnabled(True)
            if hasattr(self.ui, 'disconnect_btn'):
                self.ui.disconnect_btn.setEnabled(False)
            
            # SPI 관련 버튼들 - 초기에는 비활성화 (연결 후 활성화)
            spi_buttons = ['write_btn', 'write_all_btn', 'read_btn', 'read_all_btn']
            for btn_name in spi_buttons:
                if hasattr(self.ui, btn_name):
                    btn = getattr(self.ui, btn_name)
                    btn.setEnabled(False)  # 연결되기 전에는 비활성화
                    
            print("✅ 초기 UI 상태 설정 완료")
            
        except Exception as e:
            print(f"❌ 초기 UI 상태 설정 오류: {e}")

    def create_bit_buttons(self):
        """32개의 비트 버튼을 16x2 배열로 동적으로 생성합니다."""
        # 기존 버튼들 제거
        self.bit_buttons.clear()
        
        # 메인 수직 레이아웃 생성
        main_layout = QVBoxLayout()
        
        # 그리드 레이아웃 생성 (2행 x 16열)
        grid_layout = QGridLayout()
        grid_layout.setSpacing(2)
        
        # 0번 비트부터 31번 비트까지 (배열 인덱스와 비트 번호를 일치시킴)
        for i in range(32):
            button = QPushButton(f"{i % 10}")  # 비트 번호의 마지막 자리 표시
            button.setMinimumSize(30, 30)
            button.setMaximumSize(30, 30)
            button.setCheckable(True)
            button.setToolTip(f"Bit {i}")
            button.setStyleSheet("""
                QPushButton {
                    background-color: #f0f0f0;
                    border: 2px solid #ccc;
                    font-weight: bold;
                    font-size: 10px;
                }
                QPushButton:checked {
                    background-color: #4CAF50;
                    color: white;
                    border: 2px solid #45a049;
                }
                QPushButton:hover {
                    background-color: #e0e0e0;
                }
                QPushButton:checked:hover {
                    background-color: #45a049;
                }
            """)
            
            # 클릭 이벤트 연결 (배열 인덱스와 비트 번호가 동일)
            button.clicked.connect(lambda checked, bit_num=i: self.on_bit_button_clicked(bit_num, checked))
            
            self.bit_buttons.append(button)
            
            # 그리드에 배치: 화면 표시는 MSB부터 LSB 순서 (31→0)
            # 상위 16비트(31~16)는 0행, 하위 16비트(15~0)는 1행
            row = 0 if i >= 16 else 1
            col = (31 - i) % 16  # 31번이 0열, 0번이 15열
            grid_layout.addWidget(button, row, col)
        
        # 비트 번호 라벨 추가
        bit_labels_layout = QHBoxLayout()
        for i in range(31, -1, -1):
            if i >= 16:  # 상위 16비트
                label = QLabel(f"{i}")
                label.setAlignment(Qt.AlignCenter)
                label.setStyleSheet("font-size: 8px; color: #666;")
                bit_labels_layout.addWidget(label)
        
        bit_labels_layout2 = QHBoxLayout()
        for i in range(15, -1, -1):  # 하위 16비트
            label = QLabel(f"{i}")
            label.setAlignment(Qt.AlignCenter)
            label.setStyleSheet("font-size: 8px; color: #666;")
            bit_labels_layout2.addWidget(label)
        
        # 메인 레이아웃에 추가
        main_layout.addLayout(bit_labels_layout)
        main_layout.addLayout(grid_layout)
        main_layout.addLayout(bit_labels_layout2)
        
        # 레이아웃을 bit_buttons_widget에 설정
        self.ui.bit_buttons_widget.setLayout(main_layout)

    def connect_signals(self):
        """시그널과 슬롯을 연결합니다."""
        # 트리 위젯 클릭 이벤트
        self.ui.tree_widget.itemClicked.connect(self.on_item_clicked)
        # 선택 변경 이벤트도 추가
        self.ui.tree_widget.itemSelectionChanged.connect(self.on_selection_changed)
        
        # 메뉴 액션 연결
        self.ui.action_open_excel.triggered.connect(self.open_excel_file)
        self.ui.action_save_json.triggered.connect(self.save_json_file)
        self.ui.action_exit.triggered.connect(self.close)
        self.ui.action_expand_all.triggered.connect(self.ui.tree_widget.expandAll)
        self.ui.action_collapse_all.triggered.connect(self.ui.tree_widget.collapseAll)
        self.ui.action_about.triggered.connect(self.show_about)
        
        # FT2232H 연결 버튼들
        print("🔗 FT2232H 버튼 연결 중...")
        self.ui.connect_btn.clicked.connect(self.connect_ft2232h)
        self.ui.disconnect_btn.clicked.connect(self.disconnect_ft2232h)
        print("✅ FT2232H 버튼 연결 완료")
        
        # SPI 통신 버튼들
        print("🔗 SPI 버튼 연결 중...")
        self.ui.write_btn.clicked.connect(self.write_register)
        self.ui.write_all_btn.clicked.connect(self.write_all_registers)
        self.ui.read_btn.clicked.connect(self.read_register)
        self.ui.read_all_btn.clicked.connect(self.read_all_registers)
        print("✅ SPI 버튼 연결 완료")
        
        # 16진수 값 변경 이벤트
        self.ui.hex_value_spinbox.valueChanged.connect(self.on_hex_value_changed)
        
        # 로그 지우기 버튼
        self.ui.clear_log_btn.clicked.connect(self.clear_log)
    
    # ========== 공용 유틸리티 함수들 ==========
    
    def parse_bit_range(self, bit_range_str):
        """비트 범위 문자열을 파싱하여 (upper_bit, lower_bit) 튜플 반환"""
        try:
            bit_range_str = str(bit_range_str).strip()
            if ':' in bit_range_str:
                parts = bit_range_str.split(':')
                upper_bit = int(parts[0])
                lower_bit = int(parts[1])
            else:
                # 단일 비트인 경우
                upper_bit = lower_bit = int(bit_range_str)
            return upper_bit, lower_bit
        except (ValueError, AttributeError) as e:
            print(f"⚠️ 비트 범위 파싱 오류 ({bit_range_str}): {e}")
            return 0, 0
    
    def extract_field_value_from_register(self, register_value, upper_bit, lower_bit):
        """레지스터 값에서 특정 비트 범위의 필드 값을 추출"""
        if upper_bit >= lower_bit and upper_bit <= 31 and lower_bit >= 0:
            mask = ((1 << (upper_bit - lower_bit + 1)) - 1) << lower_bit
            field_value = (register_value & mask) >> lower_bit
            return field_value
        return 0
    
    def insert_field_value_to_register(self, register_value, field_value, upper_bit, lower_bit):
        """레지스터 값의 특정 비트 범위에 필드 값을 삽입"""
        if upper_bit >= lower_bit and upper_bit <= 31 and lower_bit >= 0:
            # 기존 해당 비트 범위를 0으로 클리어
            mask = ((1 << (upper_bit - lower_bit + 1)) - 1) << lower_bit
            register_value &= ~mask
            
            # 새로운 필드 값을 해당 위치에 삽입
            field_mask = (1 << (upper_bit - lower_bit + 1)) - 1
            shifted_value = (field_value & field_mask) << lower_bit
            register_value |= shifted_value
            
            return register_value & 0xFFFFFFFF  # 32비트 마스크
        return register_value
    
    def calculate_register_value_from_fields(self, fields):
        """필드들의 값으로부터 전체 레지스터 값을 계산"""
        total_value = 0
        
        for field in fields:
            try:
                # 기본값 파싱
                default_str = str(field.get('default_value', '0')).strip()
                if default_str.startswith('0x') or default_str.startswith('0X'):
                    field_value = int(default_str, 16)
                else:
                    field_value = int(default_str)
                
                # 비트 범위 파싱
                bit_range = field.get('bit_range', '0')
                upper_bit, lower_bit = self.parse_bit_range(bit_range)
                
                # 비트 범위가 설정되지 않았으면 기존 방식으로 가져오기
                if upper_bit == 0 and lower_bit == 0:
                    upper_bit = field.get('upper_bit', 0)
                    lower_bit = field.get('lower_bit', 0)
                
                # 필드 값이 0이 아닌 경우에만 레지스터에 값 설정
                if field_value != 0:
                    total_value = self.insert_field_value_to_register(
                        total_value, field_value, upper_bit, lower_bit
                    )
                    print(f"    🔸 필드 '{field.get('name', 'unknown')}': 값={field_value}, 비트={upper_bit}:{lower_bit}")
                else:
                    print(f"    🔸 필드 '{field.get('name', 'unknown')}': 값={field_value} (0이므로 스킵)")
                                
            except Exception as e:
                print(f"    ❌ 필드 처리 오류 '{field.get('name', 'unknown')}': {e}")
                continue
        
        final_value = total_value & 0xFFFFFFFF  # 32비트 마스크
        return final_value
    
    def calculate_register_value_from_tree(self):
        """현재 Tree 위젯에 표시된 필드 값들로부터 전체 레지스터 값을 계산"""
        if not self.current_register:
            return 0
            
        total_value = 0
        print(f"🌳 Tree 기반 레지스터 값 계산 시작 (레지스터: {self.current_register})")
        
        try:
            # Tree에서 현재 레지스터의 모든 필드 항목 찾기
            root = self.ui.tree_widget.invisibleRootItem()
            print(f"🔍 Tree 루트에서 {root.childCount()}개 시트 항목 발견")
            
            for i in range(root.childCount()):
                sheet_item = root.child(i)
                print(f"📋 시트 '{sheet_item.text(0)}'에서 {sheet_item.childCount()}개 레지스터 확인")
                
                for j in range(sheet_item.childCount()):
                    reg_item = sheet_item.child(j)
                    reg_text = reg_item.text(0)
                    reg_data = reg_item.data(0, Qt.UserRole)
                    
                    print(f"    🔍 레지스터: {reg_text}")
                    
                    # UserRole 데이터로 현재 레지스터 확인
                    if reg_data and reg_data.get('address') == self.current_register:
                        print(f"🎯 대상 레지스터 발견: {reg_text} (주소: {self.current_register})")
                        
                        # 해당 레지스터의 모든 필드 순회
                        print(f"📋 레지스터에 {reg_item.childCount()}개 필드 발견")
                        for k in range(reg_item.childCount()):
                            field_item = reg_item.child(k)
                            field_text = field_item.text(0)
                            field_data = field_item.data(0, Qt.UserRole)
                            
                            print(f"    🔍 필드 항목: '{field_text}'")
                            
                            if field_data and field_data.get('type') == 'field':
                                field_name = field_data.get('name', '')
                                bit_range = field_data.get('bit_range', '')
                                
                                # 필드 텍스트에서 값 추출 (예: "EN_VCM [15:15] = 1")
                                try:
                                    if ' = ' in field_text:
                                        field_value_text = field_text.split(' = ')[-1].strip()
                                        
                                        if field_value_text.startswith('0x') or field_value_text.startswith('0X'):
                                            field_value = int(field_value_text, 16)
                                        else:
                                            field_value = int(field_value_text)
                                        
                                        print(f"    🔢 필드 '{field_name}': Tree값={field_value}, 비트범위={bit_range}")
                                        
                                        # 비트 범위 파싱
                                        upper_bit, lower_bit = self.parse_bit_range(bit_range)
                                        
                                        # 모든 필드 값을 포함 (0이어도 처리)
                                        total_value = self.insert_field_value_to_register(
                                            total_value, field_value, upper_bit, lower_bit
                                        )
                                        print(f"    ✅ 필드 '{field_name}': 값={field_value}, 비트={upper_bit}:{lower_bit}, 누적값=0x{total_value:08X}")
                                    else:
                                        print(f"    ⚠️ 필드 텍스트에서 값을 찾을 수 없음: '{field_text}'")
                                        
                                except Exception as e:
                                    print(f"    ❌ 필드 '{field_name}' 값 파싱 오류: {e}")
                                    continue
                            else:
                                print(f"    ⚠️ 필드 데이터가 없거나 타입이 'field'가 아님")
                        
                        break
        
        except Exception as e:
            print(f"❌ Tree 기반 레지스터 값 계산 오류: {e}")
            import traceback
            traceback.print_exc()
            return 0
        
        # 대상 레지스터를 찾지 못한 경우
        if total_value == 0:
            print(f"⚠️ 레지스터 '{self.current_register}'를 Tree에서 찾을 수 없거나 모든 필드 값이 0임")
            print(f"🔄 대안: 비트 버튼 상태로 직접 계산")
            
            # 비트 버튼 상태로 직접 계산
            for i in range(32):
                if i < len(self.bit_buttons):
                    btn = self.bit_buttons[i]
                    if btn.isChecked():
                        total_value |= (1 << i)
            
            print(f"🔢 비트 버튼 기반 계산 결과: 0x{total_value:08X} ({total_value})")
        
        final_value = total_value & 0xFFFFFFFF  # 32비트 마스크
        print(f"🔢 Tree 기반 최종 레지스터 값: 0x{final_value:08X} ({final_value})")
        return final_value
    
    def get_field_data(self, field_name):
        """현재 레지스터에서 특정 필드의 데이터를 찾아 반환"""
        if not self.current_register or not hasattr(self, 'register_data'):
            return None
            
        try:
            for sheet_name, registers in self.register_data.items():
                for reg_addr, reg_info in registers.items():
                    if reg_addr == self.current_register:
                        fields = reg_info.get('fields', [])
                        for field in fields:
                            if field.get('name') == field_name:
                                return field
            return None
        except Exception as e:
            print(f"❌ 필드 데이터 검색 오류: {e}")
            return None
    
    def reset_all_bits_to_zero(self):
        """모든 비트 버튼을 0으로 초기화합니다."""
        try:
            print("🔄 모든 비트 버튼을 0으로 초기화 중...")
            
            if not self.bit_buttons or len(self.bit_buttons) != 32:
                print(f"❌ 비트 버튼 배열 문제: {len(self.bit_buttons) if self.bit_buttons else 0}")
                return
            
            # 모든 비트 버튼을 0으로 설정
            for i, button in enumerate(self.bit_buttons):
                button.blockSignals(True)
                button.setChecked(False)
                button.setText("0")
                button.blockSignals(False)
            
            # 전역 상태 초기화 (현재 선택된 레지스터가 있는 경우)
            if self.reg_addr:
                self.update_global_register_state(self.reg_addr, 0, "reset_bits")
            
            # SpinBox 값을 0으로 설정
            if hasattr(self.ui, 'hex_value_spinbox'):
                self.ui.hex_value_spinbox.blockSignals(True)
                self.ui.hex_value_spinbox.setValue(0)
                self.ui.hex_value_spinbox.blockSignals(False)
            
            # DEC 표시를 0으로 업데이트
            self.update_dec_display(0)
            
            # Tree 표시 값 업데이트
            if self.current_register:
                self.update_tree_display_values(0)
            
            print("✅ 모든 비트 버튼이 0으로 초기화됨")
            
        except Exception as e:
            print(f"❌ 비트 버튼 초기화 오류: {e}")
            import traceback
            traceback.print_exc()
    
    # ========== 메인 기능 함수들 ==========

    def on_bit_button_clicked(self, bit_index, checked):
        """비트 버튼 클릭 이벤트 처리 - 완전 안전 버전 (spinbox 제거)"""
        try:
            print(f"🔧 비트 {bit_index} 클릭: {'ON' if checked else 'OFF'}")
            
            # 입력 검증
            if not (0 <= bit_index <= 31):
                print(f"❌ 잘못된 비트 인덱스: {bit_index}")
                return
                
            if not self.bit_buttons or len(self.bit_buttons) != 32:
                print(f"❌ 비트 버튼 배열 오류: {len(self.bit_buttons) if self.bit_buttons else 0}")
                return
            
            # 버튼 텍스트 업데이트 (직접 인덱스 사용)
            if 0 <= bit_index < len(self.bit_buttons):
                button = self.bit_buttons[bit_index]
                button.setText("1" if checked else "0")
            
            # 현재 선택된 필드가 있는 경우 필드 범위 내에서만 계산
            if self.current_field_data:
                bit_range = self.current_field_data.get('bit_range', '')
                upper_bit, lower_bit = self.parse_bit_range(bit_range)
                
                print(f"🎯 필드 선택됨: {self.current_field} [{bit_range}]")
                
                # 클릭된 비트가 필드 범위 내에 있는지 확인
                if lower_bit <= bit_index <= upper_bit:
                    # 필드 범위 내의 비트들만으로 필드 값 계산
                    field_value = 0
                    for bit_pos in range(lower_bit, upper_bit + 1):
                        btn_idx = bit_pos
                        if 0 <= btn_idx < len(self.bit_buttons):
                            btn = self.bit_buttons[btn_idx]
                            if btn.isChecked():
                                # 필드 내에서의 비트 위치 (lower_bit가 0번 위치)
                                field_bit_pos = bit_pos - lower_bit
                                field_value |= (1 << field_bit_pos)
                    
                    print(f"🔢 필드 값 계산: {field_value} (비트 범위: {upper_bit}:{lower_bit})")
                    
                    # 필드 선택 시에는 필드 값만 사용 (다른 비트는 무관)
                    final_value = field_value
                    
                    print(f"🔍 필드 선택 상태: 필드 값만 사용 0x{final_value:08X} ({final_value})")
                else:
                    print(f"⚠️ 클릭된 비트 {bit_index}가 필드 범위 [{upper_bit}:{lower_bit}] 밖에 있음")
                    return
            else:
                # 필드가 선택되지 않은 경우 전체 32비트 값 계산
                calculated_value = 0
                for i in range(32):
                    if i < len(self.bit_buttons):
                        btn = self.bit_buttons[i]
                        if btn.isChecked():
                            bit_position = i  # 직접 인덱스 사용
                            if 0 <= bit_position <= 31:
                                calculated_value |= (1 << bit_position)
                
                final_value = calculated_value & 0xFFFFFFFF
            
            print(f"🔢 계산된 값: {final_value} (0x{final_value:08X})")
            
            # 전역 reg_data 업데이트
            if self.reg_addr:
                self.update_global_register_state(self.reg_addr, final_value, "bit_button_click")
            
            # SpinBox 업데이트 
            if hasattr(self, 'ui') and hasattr(self.ui, 'hex_value_spinbox'):
                # unsigned int를 signed int로 변환
                if final_value > 2147483647:
                    signed_value = final_value - 4294967296
                else:
                    signed_value = final_value
                
                if self.current_field_data:
                    # 필드가 선택된 상태에서는 시그널 차단하여 무한 루프 방지
                    self.ui.hex_value_spinbox.blockSignals(True)
                    self.ui.hex_value_spinbox.setValue(signed_value)
                    self.ui.hex_value_spinbox.blockSignals(False)
                    
                    # 직접 DEC와 Tree 업데이트 (비트 버튼은 이미 클릭으로 업데이트됨)
                    self.update_dec_display(final_value)
                    self.update_tree_display_values(final_value)
                    
                    # 디버깅: 실제 필드 값 검증
                    bit_range = self.current_field_data.get('bit_range', '')
                    upper_bit, lower_bit = self.parse_bit_range(bit_range)
                    extracted_field_value = self.extract_field_value_from_register(final_value, upper_bit, lower_bit)
                    print(f"🔍 디버깅 - 필드: {self.current_field}, 범위: {bit_range}")
                    print(f"🔍 전체 레지스터 값: 0x{final_value:08X} ({final_value})")
                    print(f"🔍 추출된 필드 값: {extracted_field_value}")
                    
                    print(f"✅ 필드 선택 상태 - 직접 업데이트: {final_value} (0x{final_value:08X})")
                else:
                    # 필드가 선택되지 않은 상태에서는 정상 시그널 발생
                    self.ui.hex_value_spinbox.setValue(signed_value)
                    print(f"✅ 비트 버튼 -> SpinBox 업데이트: {final_value} (0x{final_value:08X})")
            else:
                # SpinBox가 없으면 직접 DEC 업데이트
                self.update_dec_display(final_value)
            
        except Exception as e:
            print(f"❌ 비트 버튼 클릭 처리 중 예외: {e}")
            import traceback
            traceback.print_exc()
    
    def update_hex_display(self, value):
        """HEX 표시 업데이트 (unsigned to signed 변환)"""
        try:
            if hasattr(self, 'ui') and hasattr(self.ui, 'hex_value_spinbox'):
                hex_widget = self.ui.hex_value_spinbox
                if hex_widget is not None:
                    # unsigned int를 signed int로 변환
                    if value > 2147483647:
                        signed_value = value - 4294967296
                    else:
                        signed_value = value
                        
                    hex_widget.blockSignals(True)
                    hex_widget.setValue(signed_value)
                    hex_widget.blockSignals(False)
                    print(f"✅ HEX SpinBox 업데이트: UInt32={value} (0x{value:08X}) -> Signed={signed_value}")
        except Exception as e:
            print(f"❌ HEX 표시 업데이트 오류: {e}")
    
    def update_dec_display(self, value):
        """DEC 표시 업데이트 (안전한 방법)"""
        try:
            if hasattr(self, 'ui') and hasattr(self.ui, 'dec_value_display'):
                dec_widget = self.ui.dec_value_display
                if dec_widget is not None:
                    dec_text = str(value)
                    dec_widget.setText(dec_text)
                    print(f"✅ DEC 업데이트: {dec_text}")
        except Exception as e:
            print(f"❌ DEC 표시 업데이트 오류: {e}")

    def on_hex_value_changed(self, value):
        """QSpinBox 값 변경 이벤트 처리 - signed to unsigned 변환"""
        try:
            # UI 업데이트 중이면 처리하지 않음 (무한 루프 방지)
            if self._updating_ui:
                return
                
            # signed int를 unsigned int로 변환
            if value < 0:
                uint32_value = value + 4294967296  # 2^32
            else:
                uint32_value = value
                
            print(f"🔢 SpinBox 값 변경: {value} -> UInt32: {uint32_value} (0x{uint32_value:08X})")
            
            # 필드가 선택된 경우: 필드 범위 내에서만 값 적용
            if self.current_field_data:
                bit_range = self.current_field_data.get('bit_range', '')
                upper_bit, lower_bit = self.parse_bit_range(bit_range)
                
                # 현재 전체 비트 버튼 상태 확인
                current_full_value = 0
                for i in range(32):
                    if i < len(self.bit_buttons):
                        btn = self.bit_buttons[i]
                        if btn.isChecked():
                            current_full_value |= (1 << i)
                
                # 선택된 필드 범위의 기존 값 제거
                mask = ((1 << (upper_bit - lower_bit + 1)) - 1) << lower_bit
                current_full_value &= ~mask
                
                # 새로운 필드 값을 해당 비트 범위에 삽입
                field_value = uint32_value & ((1 << (upper_bit - lower_bit + 1)) - 1)
                final_value = current_full_value | (field_value << lower_bit)
                
                print(f"🎯 필드 '{self.current_field}' 범위 [{bit_range}]에만 값 적용")
                print(f"   기존 전체 값: 0x{current_full_value:08X}")
                print(f"   새 필드 값: {field_value}")
                print(f"   최종 값: 0x{final_value:08X}")
                
                # 비트 버튼들 업데이트 (최종 값으로)
                self.update_bit_buttons_from_value(final_value)
                
                # DEC 표시 업데이트 (최종 값으로)
                self.update_dec_display(final_value)
                
                # Tree의 표시 값 업데이트 (최종 값으로)
                self.update_tree_display_values(final_value)
                
                # 전역 reg_data 업데이트
                if self.reg_addr:
                    self.update_global_register_state(self.reg_addr, final_value, "spinbox_field_change")
            else:
                # 필드가 선택되지 않은 경우: 전체 값 적용
                print(f"📊 전체 레지스터 값 적용: 0x{uint32_value:08X}")
                
                # 비트 버튼들 업데이트
                self.update_bit_buttons_from_value(uint32_value)
                
                # DEC 표시 업데이트
                self.update_dec_display(uint32_value)
                
                # Tree의 표시 값 업데이트
                self.update_tree_display_values(uint32_value)
                
                # 전역 reg_data 업데이트
                if self.reg_addr:
                    self.update_global_register_state(self.reg_addr, uint32_value, "spinbox_full_change")
            
            print(f"✅ HEX 처리 완료: 0x{value:08X}")
            
        except Exception as e:
            print(f"❌ HEX 값 변경 처리 중 오류: {e}")
            import traceback
            traceback.print_exc()
    
    def update_bit_buttons_from_value(self, value):
        """값에서 비트 버튼들을 안전하게 업데이트 (필드 선택 상태 유지)"""
        try:
            if not self.bit_buttons or len(self.bit_buttons) != 32:
                print(f"❌ 비트 버튼 배열 문제: {len(self.bit_buttons) if self.bit_buttons else 0}")
                return
            
            # 현재 강조된 필드의 비트 범위 저장 (있다면)
            highlighted_bits = set()
            if self.current_field_data:
                bit_range = self.current_field_data.get('bit_range', '')
                upper_bit, lower_bit = self.parse_bit_range(bit_range)
                highlighted_bits = set(range(lower_bit, upper_bit + 1))
                
            for i in range(32):
                if i < len(self.bit_buttons):
                    button = self.bit_buttons[i]
                    bit_value = (value >> i) & 1  # i번째 비트 값 추출
                    
                    button.blockSignals(True)
                    button.setChecked(bool(bit_value))
                    button.setText("1" if bit_value else "0")
                    button.blockSignals(False)
                    
                    # 강조된 비트가 아닌 경우에만 기본 스타일 적용
                    if i not in highlighted_bits:
                        # 기본 스타일 유지 (강조 표시되지 않은 버튼들)
                        pass  # 이미 기본 스타일이 적용되어 있음
                    
        except Exception as e:
            print(f"❌ 비트 버튼 업데이트 오류: {e}")

    def connect_ft2232h(self):
        """FT2232H SPI 연결"""
        print("🔗 FT2232H 연결 버튼 클릭됨")
        
        if not PYFTDI_AVAILABLE:
            QMessageBox.warning(self, "라이브러리 없음", "pyftdi 라이브러리가 필요합니다.\npip install pyftdi")
            return
            
        try:
            url = self.ui.url_edit.text()
            frequency = int(self.ui.freq_edit.text())
            
            # SPI 컨트롤러 초기화
            self.spi_controller = SpiController()
            self.spi_controller.configure(url)
            self.spi = self.spi_controller.get_port(cs=0, freq=frequency, mode=0)
            
            # UI 상태 변경
            self.ui.connect_btn.setEnabled(False)
            self.ui.disconnect_btn.setEnabled(True)
            self.ui.write_btn.setEnabled(True)
            self.ui.write_all_btn.setEnabled(True)
            self.ui.read_btn.setEnabled(True)
            self.ui.read_all_btn.setEnabled(True)
            
            print("✅ SPI 버튼들 활성화됨")
            self.log_message(f"✅ FT2232H 연결 성공: {url} @ {frequency}Hz")
            self.statusBar().showMessage("FT2232H 연결됨")
            
        except Exception as e:
            QMessageBox.critical(self, "연결 오류", f"FT2232H 연결 실패:\n{str(e)}")
            self.log_message(f"❌ FT2232H 연결 실패: {str(e)}")

    def disconnect_ft2232h(self):
        """FT2232H SPI 연결 해제"""
        print("🔌 FT2232H 연결 해제 버튼 클릭됨")
        
        try:
            if self.spi_controller:
                self.spi_controller.close()
                self.spi_controller = None
                self.spi = None
            
            # UI 상태 변경
            self.ui.connect_btn.setEnabled(True)
            self.ui.disconnect_btn.setEnabled(False)
            self.ui.write_btn.setEnabled(False)
            self.ui.write_all_btn.setEnabled(False)
            self.ui.read_btn.setEnabled(False)
            self.ui.read_all_btn.setEnabled(False)
            
            print("✅ SPI 버튼들 비활성화됨")
            self.log_message("🔌 FT2232H 연결 해제")
            self.statusBar().showMessage("연결 해제됨")
            
        except Exception as e:
            self.log_message(f"❌ 연결 해제 오류: {str(e)}")

    def write_register(self):
        """현재 선택된 레지스터에 값 쓰기 (전역 reg_data 사용)"""
        print("✍️ Write Register 버튼 클릭됨")
        
        if not self.spi or not self.reg_addr:
            QMessageBox.warning(self, "경고", "SPI가 연결되지 않았거나 레지스터가 선택되지 않았습니다.")
            return
            
        try:
            addr = int(self.reg_addr, 16)
            
            # 전역 reg_data 사용 (현재 레지스터의 실제 값)
            value = self.reg_data
            
            print(f"📊 전역 reg_data 기반 레지스터 값: 0x{value:08X} ({value})")
            print(f"🎯 레지스터 주소: {self.reg_addr}")
            
            if self.current_field and self.current_field_data:
                # 필드가 선택된 경우: 해당 필드 값 표시
                bit_range = self.current_field_data.get('bit_range', '')
                upper_bit, lower_bit = self.parse_bit_range(bit_range)
                field_value = self.extract_field_value_from_register(value, upper_bit, lower_bit)
                print(f"🎯 필드 '{self.current_field}' 선택됨:")
                print(f"   전체 레지스터 값: 0x{value:08X} ({value})")
                print(f"   필드 범위 [{bit_range}] 값: {field_value}")
            else:
                # 레지스터 전체 선택된 경우
                print(f"📊 전체 레지스터 값: 0x{value:08X} ({value})")
            
            # SPI 쓰기 명령 (RW=0, 주소 7비트 + 데이터 32비트)
            write_cmd = (addr & 0x7F)  # RW=0, 주소 7비트
            data_byte3 = (value >> 24) & 0xFF
            data_byte2 = (value >> 16) & 0xFF
            data_byte1 = (value >> 8) & 0xFF
            data_byte0 = value & 0xFF
            
            # SPI 전송 (32비트 데이터)
            response = self.spi.exchange([write_cmd, data_byte3, data_byte2, data_byte1, data_byte0])
            
            if self.current_field:
                self.log_message(f"📝 WRITE (필드 '{self.current_field}'): Addr={self.reg_addr}, Value=0x{value:08X} ({value})")
            else:
                self.log_message(f"📝 WRITE: Addr={self.reg_addr}, Value=0x{value:08X} ({value})")
            self.log_message(f"   CMD: 0x{write_cmd:02X} 0x{data_byte3:02X} 0x{data_byte2:02X} 0x{data_byte1:02X} 0x{data_byte0:02X}")
            
        except Exception as e:
            QMessageBox.critical(self, "쓰기 오류", f"레지스터 쓰기 실패:\n{str(e)}")
            self.log_message(f"❌ 쓰기 실패: {str(e)}")

    def write_all_registers(self):
        """모든 레지스터에 현재 값 쓰기"""
        print("✍️ Write All Registers 버튼 클릭됨")
        
        if not self.spi or not self.data:
            QMessageBox.warning(self, "경고", "SPI가 연결되지 않았거나 데이터가 없습니다.")
            return
            
        try:
            # QSpinBox에서 값 가져와서 unsigned로 변환
            signed_value = self.ui.hex_value_spinbox.value()
            if signed_value < 0:
                value = signed_value + 4294967296  # 2^32
            else:
                value = signed_value
            
            count = 0
            for registers in self.data.values():
                for register in registers:
                    addr = int(register['address'], 16)
                    
                    # SPI 쓰기 명령
                    write_cmd = (addr & 0x7F)
                    data_byte3 = (value >> 24) & 0xFF
                    data_byte2 = (value >> 16) & 0xFF
                    data_byte1 = (value >> 8) & 0xFF
                    data_byte0 = value & 0xFF
                    
                    response = self.spi.exchange([write_cmd, data_byte3, data_byte2, data_byte1, data_byte0])
                    count += 1
            
            self.log_message(f"📝 WRITE ALL: {count}개 레지스터에 0x{value:08X} 쓰기 완료")
            
        except Exception as e:
            QMessageBox.critical(self, "쓰기 오류", f"전체 쓰기 실패:\n{str(e)}")
            self.log_message(f"❌ 전체 쓰기 실패: {str(e)}")

    def read_register(self):
        """현재 선택된 레지스터 읽기"""
        print("📖 Read Register 버튼 클릭됨")
        
        if not self.spi or not self.current_register:
            QMessageBox.warning(self, "경고", "SPI가 연결되지 않았거나 레지스터가 선택되지 않았습니다.")
            return
            
        try:
            addr = int(self.current_register, 16)
            
            # SPI 읽기 명령 (RW=1, 주소 7비트)
            read_cmd = 0x80 | (addr & 0x7F)  # RW=1, 주소 7비트
            
            # SPI 전송 (읽기는 4바이트 수신)
            response = self.spi.exchange([read_cmd, 0x00, 0x00, 0x00, 0x00])
            
            # 응답에서 데이터 추출 (첫 바이트는 명령 에코)
            if len(response) >= 5:
                value = (response[1] << 24) | (response[2] << 16) | (response[3] << 8) | response[4]
            else:
                value = 0
                
            # UI 업데이트 (unsigned to signed 변환)
            if value > 2147483647:
                signed_value = value - 4294967296
            else:
                signed_value = value
            self.ui.hex_value_spinbox.setValue(signed_value)
            
            self.log_message(f"📖 READ: Addr={self.current_register}, Value=0x{value:08X} ({value})")
            
        except Exception as e:
            QMessageBox.critical(self, "읽기 오류", f"레지스터 읽기 실패:\n{str(e)}")
            self.log_message(f"❌ 읽기 실패: {str(e)}")

    def read_all_registers(self):
        """모든 레지스터 읽기"""
        print("📖 Read All Registers 버튼 클릭됨")
        
        if not self.spi or not self.data:
            QMessageBox.warning(self, "경고", "SPI가 연결되지 않았거나 데이터가 없습니다.")
            return
            
        try:
            count = 0
            for registers in self.data.values():
                for register in registers:
                    addr = int(register['address'], 16)
                    
                    # SPI 읽기 명령
                    read_cmd = 0x80 | (addr & 0x7F)
                    response = self.spi.exchange([read_cmd, 0x00, 0x00, 0x00, 0x00])
                    
                    if len(response) >= 5:
                        value = (response[1] << 24) | (response[2] << 16) | (response[3] << 8) | response[4]
                        self.log_message(f"📖 READ: Addr=0x{addr:02X}, Value=0x{value:08X}")
                    
                    count += 1
            
            self.log_message(f"📖 READ ALL: {count}개 레지스터 읽기 완료")
            
        except Exception as e:
            QMessageBox.critical(self, "읽기 오류", f"전체 읽기 실패:\n{str(e)}")
            self.log_message(f"❌ 전체 읽기 실패: {str(e)}")

    def log_message(self, message):
        """로그 메시지 추가"""
        self.ui.log_text.append(message)
        # 스크롤을 맨 아래로
        scrollbar = self.ui.log_text.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())

    def clear_log(self):
        """로그 지우기"""
        self.ui.log_text.clear()

    def on_item_clicked(self, item, column):
        """트리 아이템 클릭 이벤트"""
        print(f"🖱️ 트리 아이템 클릭됨: '{item.text(0)}'")
        
        # 아이템의 사용자 데이터 확인
        item_data = item.data(0, Qt.UserRole)
        print(f"📄 아이템 데이터: {item_data}")
        
        if item_data and isinstance(item_data, dict):
            print(f"🔍 데이터 타입: {item_data.get('type')}")
            
            if item_data.get('type') == 'register':
                # 레지스터 선택
                self.current_register = item_data['address']
                self.current_register_data = item_data
                
                # 전역 레지스터 상태 업데이트
                addr = item_data['address']
                data = self.calculate_register_default_value(item_data.get('fields', []))
                self.update_global_register_state(addr, data, "register_selection")
                
                # 현재 선택된 필드 해제 (레지스터 전체 선택)
                self.current_field = None
                self.current_field_data = None
                
                print(f"🎯 레지스터 선택: {self.current_register} - {item_data.get('description', '')}")
                
                # 필드들로부터 기본값 계산
                default_value = self.reg_data
                print(f"📊 계산된 기본값: 0x{default_value:08X} ({default_value})")
                
                # 레지스터 선택 시 비트 강조 해제
                self.clear_bit_highlights()
                
                # SpinBox 범위 복원 (전체 32비트)
                self.reset_spinbox_range()
                
                self.update_register_ui(item_data, default_value)
                
            elif item_data.get('type') == 'field':
                # 필드 선택 시 부모 레지스터 찾기
                print(f"🔧 필드 선택: {item_data.get('name')} [{item_data.get('bit_range')}]")
                
                parent_item = item.parent()
                if parent_item:
                    parent_data = parent_item.data(0, Qt.UserRole)
                    if parent_data and parent_data.get('type') == 'register':
                        # 부모 레지스터 자동 선택
                        self.current_register = parent_data['address']
                        self.current_register_data = parent_data
                        
                        # 전역 레지스터 상태 업데이트 (현재 값 유지)
                        addr = parent_data['address']
                        if self.reg_addr != addr:
                            # 다른 레지스터로 변경된 경우만 초기화
                            data = self.calculate_register_default_value(parent_data.get('fields', []))
                            self.update_global_register_state(addr, data, "field_selection_new_register")
                        else:
                            # 같은 레지스터 내의 필드 선택이면 현재 값 유지
                            try:
                                current_signed_value = self.ui.hex_value_spinbox.value()
                                if current_signed_value < 0:
                                    current_data = current_signed_value + 4294967296
                                else:
                                    current_data = current_signed_value
                                self.update_global_register_state(addr, current_data, "field_selection_same_register")
                            except:
                                # SpinBox 값을 가져올 수 없으면 현재 값 유지
                                print(f"📊 현재 reg_data 값 유지: 0x{self.reg_data:08X}")
                        
                        # 현재 선택된 필드 정보 저장
                        self.current_field = item_data.get('name', '')
                        self.current_field_data = item_data
                        
                        print(f"🎯 부모 레지스터 자동 선택: {self.current_register} - {parent_data.get('description', '')}")
                        print(f"🎯 현재 선택된 필드: {self.current_field}")
                        
                        # UI 업데이트 (현재 값 유지)
                        self.update_register_ui_preserve_value(parent_data, self.reg_data)
                        
                        # 선택된 필드 강조 표시
                        field_name = item_data.get('name', '')
                        bit_range = item_data.get('bit_range', '')
                        print(f"✨ 선택된 필드: {field_name} [{bit_range}]")
                        
                        # 선택된 필드의 비트 버튼들 강조 (현재 레지스터 값에서 해당 필드만 추출)
                        self.highlight_field_bits(item_data)
        else:
            print(f"⚠️ 아이템 데이터가 없거나 올바르지 않음")
                
    def update_register_ui(self, register_data, default_value):
        """레지스터 UI 업데이트"""
        try:
            # Current Register 표시 업데이트
            if hasattr(self.ui, 'current_register_label'):
                self.ui.current_register_label.setText(f"Current Register: 0x{self.current_register} - {register_data.get('description', '')}")
                print(f"✅ Current Register 라벨 업데이트됨")
            else:
                print(f"⚠️ current_register_label을 찾을 수 없음")
                
            # 다른 UI 요소들도 확인
            if hasattr(self.ui, 'current_register_addr'):
                self.ui.current_register_addr.setText(f"0x{self.current_register}")
                print(f"✅ current_register_addr 업데이트됨")
            if hasattr(self.ui, 'current_register_name'):
                self.ui.current_register_name.setText(register_data.get('description', ''))
                print(f"✅ current_register_name 업데이트됨")
                
            # QSpinBox에 값 설정 (unsigned to signed 변환)
            self._updating_ui = True  # 플래그 설정
            try:
                if default_value > 2147483647:
                    signed_value = default_value - 4294967296
                else:
                    signed_value = default_value
                self.ui.hex_value_spinbox.setValue(signed_value)
                print(f"✅ SpinBox 값 설정: {signed_value}")
            finally:
                self._updating_ui = False  # 플래그 해제
            
            # 비트 버튼들 업데이트
            self.update_bit_buttons_from_value(default_value)
            
            # DEC 표시 업데이트
            self.update_dec_display(default_value)
            print(f"✅ DEC 표시 업데이트 완료")
                    
        except Exception as e:
            print(f"❌ 값 업데이트 오류: {e}")
            import traceback
            traceback.print_exc()

    def update_register_ui_preserve_value(self, register_data, current_value):
        """레지스터 UI 업데이트 (현재 값 보존)"""
        try:
            # Current Register 표시 업데이트
            if hasattr(self.ui, 'current_register_label'):
                self.ui.current_register_label.setText(f"Current Register: 0x{self.current_register} - {register_data.get('description', '')}")
                print(f"✅ Current Register 라벨 업데이트됨")
                
            # 다른 UI 요소들도 확인
            if hasattr(self.ui, 'current_register_addr'):
                self.ui.current_register_addr.setText(f"0x{self.current_register}")
                print(f"✅ current_register_addr 업데이트됨")
            if hasattr(self.ui, 'current_register_name'):
                self.ui.current_register_name.setText(register_data.get('description', ''))
                print(f"✅ current_register_name 업데이트됨")
                
            # 현재 값이 유효하지 않으면 SpinBox 값만 업데이트 (비트 버튼은 그대로 유지)
            print(f"✅ 현재 값 보존됨: 0x{current_value:08X} ({current_value})")
                    
            # 현재 값이 유효하지 않으면 SpinBox 값만 업데이트 (비트 버튼은 그대로 유지)
            print(f"✅ 현재 값 보존됨: 0x{current_value:08X} ({current_value})")
                    
        except Exception as e:
            print(f"❌ UI 업데이트 오류: {e}")
            import traceback
            traceback.print_exc()
    
    def highlight_field_bits(self, field_data):
        """선택된 필드의 비트 버튼들을 강조 표시하고 활성화합니다."""
        try:
            # 모든 비트 버튼 강조 해제
            self.clear_bit_highlights()
            
            # 비트 범위 파싱 (공용 함수 사용)
            bit_range = field_data.get('bit_range', '')
            upper_bit, lower_bit = self.parse_bit_range(bit_range)
            
            print(f"🎯 필드 비트 범위 강조: {upper_bit}:{lower_bit}")
            
            # SpinBox 범위 제한 설정
            self.set_spinbox_range_for_field(upper_bit, lower_bit)
            
            # 모든 비트 버튼을 순회하면서 상태 설정
            for i in range(32):
                if i < len(self.bit_buttons):
                    button = self.bit_buttons[i]
                    bit_pos = i  # 직접 인덱스 사용
                    
                    if lower_bit <= bit_pos <= upper_bit:
                        # 필드 범위 내의 버튼: 강조 표시 + 활성화
                        button.setEnabled(True)
                        button.setStyleSheet("""
                            QPushButton {
                                background-color: #FFD700;
                                border: 3px solid #FFA500;
                                font-weight: bold;
                                font-size: 10px;
                                color: #000;
                            }
                            QPushButton:checked {
                                background-color: #FF8C00;
                                color: white;
                                border: 3px solid #FF6347;
                            }
                            QPushButton:hover {
                                background-color: #FFA500;
                            }
                            QPushButton:checked:hover {
                                background-color: #FF6347;
                            }
                        """)
                        print(f"✨ 비트 {bit_pos} 버튼 강조 및 활성화됨")
                    else:
                        # 필드 범위 밖의 버튼: 비활성화
                        button.setEnabled(False)
                        button.setStyleSheet("""
                            QPushButton {
                                background-color: #e0e0e0;
                                border: 2px solid #bbb;
                                font-weight: normal;
                                font-size: 10px;
                                color: #888;
                            }
                            QPushButton:checked {
                                background-color: #ccc;
                                color: #666;
                                border: 2px solid #999;
                            }
                        """)
            
            # 필드의 현재 값을 비트 버튼에 적용
            self.apply_field_value_to_bit_buttons(field_data, upper_bit, lower_bit)
            
        except Exception as e:
            print(f"❌ 필드 비트 강조 오류: {e}")
    
    def apply_field_value_to_bit_buttons(self, field_data, upper_bit, lower_bit):
        """선택된 필드의 현재 값을 비트 버튼에 적용합니다 (필드 범위 외 비트는 클리어)."""
        try:
            field_name = field_data.get('name', '')
            
            # 전역 reg_data에서 해당 필드의 현재 값 추출
            field_value = self.extract_field_value_from_register(self.reg_data, upper_bit, lower_bit)
            
            print(f"🎯 필드 '{field_name}' 현재 값: {field_value} → 비트 버튼에 적용")
            print(f"   📊 전체 레지스터 값: 0x{self.reg_data:08X}, 필드 범위: [{upper_bit}:{lower_bit}]")
            
            # UI 업데이트 플래그 설정 (무한 루프 방지)
            self._updating_ui = True
            
            # 모든 비트 버튼을 먼저 0으로 클리어 (필드 선택 시에는 필드 외 비트는 무관)
            for i in range(32):
                if i < len(self.bit_buttons):
                    button = self.bit_buttons[i]
                    if not (lower_bit <= i <= upper_bit):
                        # 필드 범위 외의 비트는 0으로 클리어
                        button.setChecked(False)
                        button.setText("0")
            
            # 필드 범위의 각 비트에 대해 버튼 상태 설정
            for bit_pos in range(lower_bit, upper_bit + 1):
                if bit_pos < len(self.bit_buttons):
                    button = self.bit_buttons[bit_pos]
                    
                    # 필드 값에서 해당 비트 위치의 값 추출
                    field_bit_pos = bit_pos - lower_bit  # 필드 내에서의 비트 위치
                    bit_value = (field_value >> field_bit_pos) & 1
                    
                    # 버튼 상태 설정
                    button.setChecked(bit_value == 1)
                    button.setText("1" if bit_value == 1 else "0")
                    
                    print(f"  📌 비트 {bit_pos}: {bit_value}")
            
            # UI 업데이트 플래그 해제
            self._updating_ui = False
            
            # SpinBox 값은 필드 값만 표시 (전체 레지스터 값이 아님)
            if hasattr(self.ui, 'hex_value_spinbox'):
                if field_value > 2147483647:
                    signed_value = field_value - 4294967296
                else:
                    signed_value = field_value
                
                self.ui.hex_value_spinbox.blockSignals(True)
                self.ui.hex_value_spinbox.setValue(signed_value)
                self.ui.hex_value_spinbox.blockSignals(False)
            
            # DEC 표시 업데이트 (필드 값만)
            self.update_dec_display(field_value)
            
            # reg_data를 필드 값만으로 업데이트 (필드 선택 시에는 필드 값만 유효)
            if self.reg_addr:
                self.update_global_register_state(self.reg_addr, field_value, "field_value_apply")
            
            print(f"✅ 필드 값 {field_value} (0x{field_value:X})가 비트 버튼에 적용됨")
            print(f"📊 reg_data 업데이트: 0x{self.reg_data:08X} (필드 값만)")
            
        except Exception as e:
            print(f"❌ 필드 값 적용 오류: {e}")
    
    
    def extract_field_value_from_register(self, register_value, upper_bit, lower_bit):
        """레지스터 값에서 특정 필드의 값을 추출합니다."""
        try:
            # 비트 범위만큼 마스크 생성
            bit_count = upper_bit - lower_bit + 1
            mask = (1 << bit_count) - 1
            
            # 레지스터 값에서 해당 필드 추출
            field_value = (register_value >> lower_bit) & mask
            
            return field_value
            
        except Exception as e:
            print(f"❌ 필드 값 추출 오류: {e}")
            return 0
    
    def update_global_register_state(self, addr, data, source="unknown"):
        """전역 레지스터 상태를 업데이트합니다."""
        try:
            self.reg_addr = addr
            self.reg_data = data
            print(f"📊 전역 상태 업데이트 ({source}): reg_addr={self.reg_addr}, reg_data=0x{self.reg_data:08X} ({self.reg_data})")
            
        except Exception as e:
            print(f"❌ 전역 상태 업데이트 오류: {e}")
    
    def get_field_value_from_tree(self, field_name):
        """Tree에서 특정 필드의 현재 값을 가져옵니다."""
        try:
            if not self.current_register:
                return None
                
            root = self.ui.tree_widget.invisibleRootItem()
            
            for i in range(root.childCount()):
                sheet_item = root.child(i)
                for j in range(sheet_item.childCount()):
                    reg_item = sheet_item.child(j)
                    reg_data = reg_item.data(0, Qt.UserRole)
                    
                    if reg_data and reg_data.get('address') == self.current_register:
                        # 해당 레지스터의 필드들 확인
                        for k in range(reg_item.childCount()):
                            field_item = reg_item.child(k)
                            field_text = field_item.text(0)
                            field_data = field_item.data(0, Qt.UserRole)
                            
                            if field_data and field_data.get('name') == field_name:
                                # 필드 텍스트에서 값 추출 (예: "EN_VCM [15:15] = 1")
                                if ' = ' in field_text:
                                    value_text = field_text.split(' = ')[-1].strip()
                                    
                                    if value_text.startswith('0x') or value_text.startswith('0X'):
                                        return int(value_text, 16)
                                    else:
                                        return int(value_text)
                                return 0  # 값이 없으면 0 반환
            
            return None
            
        except Exception as e:
            print(f"❌ Tree에서 필드 값 조회 오류: {e}")
            return None
    
    def set_spinbox_range_for_field(self, upper_bit, lower_bit):
        """선택된 필드의 비트 범위에 따라 SpinBox 범위를 제한합니다."""
        try:
            # 필드 비트 수 계산
            bit_count = upper_bit - lower_bit + 1
            
            # 최대값 계산 (2^bit_count - 1)
            max_value = (1 << bit_count) - 1
            min_value = 0
            
            print(f"📊 필드 범위 제한: {bit_count}비트 → 0 ~ {max_value} (0x{max_value:X})")
            
            # SpinBox 범위 설정 (unsigned 값 기준)
            # signed 범위로 변환: 양수는 그대로, 음수는 2^32를 빼서 표현
            if max_value <= 2147483647:
                # 양수 범위
                signed_min = min_value
                signed_max = max_value
            else:
                # 큰 값은 signed로 변환 불가하므로 32비트 전체 범위 유지
                signed_min = -2147483648
                signed_max = 2147483647
            
            # SpinBox 범위 적용
            if hasattr(self.ui, 'hex_value_spinbox'):
                self.ui.hex_value_spinbox.setRange(signed_min, signed_max)
                print(f"✅ SpinBox 범위 설정: {signed_min} ~ {signed_max}")
                
                # 현재 값이 범위를 벗어나면 0으로 초기화
                current_value = self.ui.hex_value_spinbox.value()
                if current_value < signed_min or current_value > signed_max:
                    self.ui.hex_value_spinbox.setValue(0)
                    print(f"⚠️ 현재 값이 범위를 벗어나서 0으로 초기화됨")
            
        except Exception as e:
            print(f"❌ SpinBox 범위 설정 오류: {e}")
    
    def reset_spinbox_range(self):
        """SpinBox 범위를 전체 32비트로 복원합니다."""
        try:
            if hasattr(self.ui, 'hex_value_spinbox'):
                # 32비트 전체 범위로 복원 (signed int 범위)
                self.ui.hex_value_spinbox.setRange(-2147483648, 2147483647)
                print(f"✅ SpinBox 범위 복원: 32비트 전체 범위")
        except Exception as e:
            print(f"❌ SpinBox 범위 복원 오류: {e}")
    
    def clear_bit_highlights(self):
        """모든 비트 버튼의 강조를 해제하고 모든 버튼을 활성화합니다."""
        try:
            for button in self.bit_buttons:
                # 모든 버튼 활성화
                button.setEnabled(True)
                # 기본 스타일로 복원
                button.setStyleSheet("""
                    QPushButton {
                        background-color: #f0f0f0;
                        border: 2px solid #ccc;
                        font-weight: bold;
                        font-size: 10px;
                    }
                    QPushButton:checked {
                        background-color: #4CAF50;
                        color: white;
                        border: 2px solid #45a049;
                    }
                    QPushButton:hover {
                        background-color: #e0e0e0;
                    }
                    QPushButton:checked:hover {
                        background-color: #45a049;
                    }
                """)
        except Exception as e:
            print(f"❌ 비트 버튼 강조 해제 오류: {e}")
    
    def on_selection_changed(self):
        """트리 아이템 선택 변경 이벤트"""
        print("🖱️ 트리 선택 변경 감지됨")
        
        current_items = self.ui.tree_widget.selectedItems()
        if current_items:
            item = current_items[0]  # 첫 번째 선택된 아이템
            print(f"📋 선택된 아이템: '{item.text(0)}'")
            
            # 기존 on_item_clicked와 동일한 로직 실행
            self.on_item_clicked(item, 0)
        else:
            print("⚠️ 선택된 아이템이 없음")
    
    def calculate_register_default_value(self, fields):
        """필드들의 기본값으로부터 전체 레지스터 기본값을 계산합니다."""
        print(f"  📊 레지스터 기본값 계산 시작 (필드 수: {len(fields)})")
        
        final_value = self.calculate_register_value_from_fields(fields)
        print(f"  📊 최종 레지스터 기본값: 0x{final_value:08X} ({final_value})")
        return final_value

    def load_excel_file(self, file_path):
        """Excel 파일을 로드합니다."""
        try:
            self.data = self.load_excel(file_path)
            if self.data:
                self.build_tree()
                self.log_message(f"✅ Excel 파일 로드: {file_path}")
            else:
                self.log_message(f"❌ Excel 파일 로드 실패: {file_path}")
        except Exception as e:
            QMessageBox.critical(self, "파일 로드 오류", f"Excel 파일을 로드할 수 없습니다:\\n{str(e)}")
            self.log_message(f"❌ Excel 로드 오류: {str(e)}")

    def load_excel(self, file_path):
        """Excel 파일에서 레지스터 정보를 읽어옵니다 (개선된 병합 셀 처리)."""
        try:
            print(f"📂 Excel 파일 로딩 시작: {file_path}")
            
            # pandas로 데이터 읽기
            df = pd.read_excel(file_path, header=None)
            
            # openpyxl로 병합된 셀 정보 읽기
            wb = load_workbook(file_path, read_only=False)
            sheet = wb.active
            merged_ranges = sheet.merged_cells.ranges
            
            print(f"📊 Excel 파일 크기: {df.shape[0]}행 x {df.shape[1]}열")
            print(f"🔗 병합된 셀 범위: {len(merged_ranges)}개")
            
            # 병합된 셀 정보를 딕셔너리로 변환 (더 빠른 검색을 위해)
            merged_info = {}
            for merged_range in merged_ranges:
                min_row, min_col = merged_range.min_row - 1, merged_range.min_col - 1  # 0-based 인덱스
                max_row, max_col = merged_range.max_row - 1, merged_range.max_col - 1
                for r in range(min_row, max_row + 1):
                    for c in range(min_col, max_col + 1):
                        merged_info[(r, c)] = {
                            'min_row': min_row, 'max_row': max_row,
                            'min_col': min_col, 'max_col': max_col,
                            'is_master': (r == min_row and c == min_col)
                        }
            
            print(f"🔧 병합된 셀 정보 인덱스 구축 완료: {len(merged_info)}개 셀")
            
            # Meaning 테이블들을 찾아서 필드 의미 매핑 생성 (개선된 방법)
            field_meanings = self.extract_all_meaning_tables_improved(df)
            
            # 레지스터 데이터 파싱 (개선된 방법)
            data = {"registers": []}
            register_count = 0
            
            print(f"🔍 레지스터 검색 시작 (전체 DataFrame 스캔)")
            
            # DataFrame에서 "Addr" 키워드를 찾아 레지스터 시작점 확인
            for row_idx in range(len(df)):
                for col_idx in range(min(5, len(df.columns))):  # 첫 5열만 확인
                    cell_value = df.iat[row_idx, col_idx]
                    
                    # 디버깅: 첫 10행의 값들 출력
                    if row_idx <= 10 and col_idx <= 4:
                        print(f"   Row {row_idx}, Col {col_idx}: '{cell_value}' (type: {type(cell_value)})")
                    
                    # "Addr" 키워드를 찾아 레지스터 시작점 확인
                    if pd.notna(cell_value) and str(cell_value).strip() == "Addr":
                        addr_col = col_idx
                        print(f"\\n🎯 레지스터 발견: Row {row_idx}, Col {col_idx} (Addr 열)")
                        
                        register_data = self.parse_register_at_row_improved(df, row_idx, addr_col, merged_info, field_meanings)
                        if register_data:
                            data["registers"].append(register_data)
                            register_count += 1
                            print(f"✅ 레지스터 #{register_count} 추가됨")
                        else:
                            print(f"❌ 레지스터 파싱 실패")
                        break  # 이 행에서 Addr을 찾았으면 다음 행으로
            
            print(f"\\n📊 총 {register_count}개 레지스터 파싱 완료")
            
            # 임시로 샘플 데이터 추가 (파싱이 실패한 경우)
            if register_count == 0:
                print("⚠️ 레지스터가 발견되지 않아 샘플 데이터 추가")
                sample_register = {
                    "address": "0x00",
                    "description": "Sample Register",
                    "fields": [
                        {
                            "name": "sample_field",
                            "bit_range": "15:0",
                            "upper_bit": 15,
                            "lower_bit": 0,
                            "default_value": "0",
                            "meaning": "Sample field for testing"
                        }
                    ],
                    "default_value": 0
                }
                data["registers"].append(sample_register)
                print("✅ 샘플 레지스터 추가됨")
            
            # JSON 파일로 저장
            json_path = file_path.replace('.xlsx', '_tree.json').replace('.xls', '_tree.json')
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            print(f"✅ JSON 파일 저장 완료: {json_path}")
            print(f"📝 저장된 레지스터 수: {len(data['registers'])}")
            
            # 워크북 닫기
            wb.close()
            
            return {"Sheet1": data["registers"]}
            
        except Exception as e:
            print(f"❌ Excel 파일 로드 중 오류: {e}")
            import traceback
            traceback.print_exc()
            return None

    def extract_all_meaning_tables_improved(self, df):
        """모든 Meaning 테이블을 찾아서 필드 의미를 추출합니다 (개선된 방법)."""
        field_meanings = {}
        
        # Excel 파일 전체에서 모든 Meaning 열 찾기
        meaning_positions = []
        for row_idx in range(len(df)):
            for col_idx in range(len(df.columns)):
                cell_value = df.iat[row_idx, col_idx]
                if pd.notna(cell_value) and str(cell_value).strip() == "Meaning":
                    meaning_positions.append((row_idx, col_idx))
        
        print(f"🔍 총 {len(meaning_positions)}개의 Meaning 테이블 발견")
        
        # 각 Meaning 테이블에서 정보 수집
        for table_idx, (meaning_row, meaning_col) in enumerate(meaning_positions):
            print(f"📋 Meaning 테이블 #{table_idx + 1} 처리 중 (Row {meaning_row}, Col {meaning_col})")
            
            # Name 열은 보통 Meaning 열보다 4칸 앞에 위치
            name_col = meaning_col - 4
            if name_col < 0:
                continue
                
            # 해당 테이블의 데이터 행들 처리
            for data_row_idx in range(meaning_row + 1, len(df)):
                if data_row_idx >= len(df):
                    break
                    
                data_row = df.iloc[data_row_idx]
                
                # Name과 Meaning 값 가져오기
                name_val = data_row.iloc[name_col] if name_col < len(data_row) else None
                meaning_val = data_row.iloc[meaning_col] if meaning_col < len(data_row) else None
                
                # 유효한 데이터인지 확인
                if pd.notna(name_val) and pd.notna(meaning_val):
                    name_str = str(name_val).strip()
                    meaning_str = str(meaning_val).strip()
                    
                    # "Name" 헤더가 다시 나오면 다음 테이블 시작이므로 중단
                    if name_str == "Name":
                        break
                        
                    if name_str and meaning_str and meaning_str != "nan":
                        field_meanings[name_str] = meaning_str
                        print(f"   📝 필드 의미: {name_str} = {meaning_str}")
                
                # 빈 행이 연속으로 나오면 테이블 끝
                elif pd.isna(name_val) and pd.isna(meaning_val):
                    # 다음 몇 행도 확인해서 정말 끝인지 체크
                    empty_count = 0
                    for check_row in range(data_row_idx, min(data_row_idx + 3, len(df))):
                        check_data = df.iloc[check_row]
                        if pd.isna(check_data.iloc[name_col]) and pd.isna(check_data.iloc[meaning_col]):
                            empty_count += 1
                    if empty_count >= 2:  # 2행 이상 비어있으면 테이블 끝
                        break
        
        return field_meanings

    def parse_register_at_row_improved(self, df, register_row, addr_col, merged_info, field_meanings):
        """특정 행에서 레지스터 정보를 파싱합니다 (개선된 병합 셀 처리)."""
        try:
            # 주소 값과 레지스터 이름 읽기
            addr_value = df.iat[register_row, addr_col + 1] if addr_col + 1 < len(df.columns) else None
            reg_name = df.iat[register_row, addr_col + 2] if addr_col + 2 < len(df.columns) else None
            
            if pd.isna(addr_value):
                return None
            
            address = str(addr_value).strip()
            description = str(reg_name).strip() if pd.notna(reg_name) else ""
            
            print(f"   📍 주소: {address}, 설명: {description}")
            
            # Bit, Name, Default 행 찾기
            bit_row = register_row + 1
            name_row = register_row + 2  
            default_row = register_row + 3
            
            print(f"   📋 Bit행: {bit_row}, Name행: {name_row}, Default행: {default_row}")
            
            # 필드들 파싱 (addr_col+1부터 시작 - 비트 번호들)
            fields = []
            processed_merged_fields = set()  # 이미 처리된 병합 필드 추적
            
            # 비트 15부터 0까지 순서대로 처리 (Excel에서 왼쪽부터 오른쪽으로)
            for bit_col in range(addr_col + 1, min(addr_col + 17, len(df.columns))):  # 16비트까지
                if bit_row >= len(df) or name_row >= len(df) or default_row >= len(df):
                    break
                
                # 비트 번호 가져오기
                bit_num_cell = df.iat[bit_row, bit_col] if bit_col < len(df.columns) else None
                name_cell = df.iat[name_row, bit_col] if bit_col < len(df.columns) else None
                default_cell = df.iat[default_row, bit_col] if bit_col < len(df.columns) else None
                
                if pd.isna(bit_num_cell):
                    continue
                    
                try:
                    bit_num = int(bit_num_cell)
                except:
                    continue
                
                # Name 셀 처리 (병합된 셀 고려)
                field_name = ""
                if pd.notna(name_cell):
                    field_name = str(name_cell).strip()
                
                # 필드명 정리 (더 나은 처리)
                clean_name = ""
                if field_name and field_name not in ["nan", ""]:
                    # Verilog 스타일의 1'b0, 1'b1 처리
                    if field_name.startswith("1'b"):
                        # 1'b0 -> BIT0, 1'b1 -> BIT1 등으로 변환
                        bit_value = field_name.replace("1'b", "")
                        clean_name = f"BIT{bit_value}"
                    else:
                        # 일반적인 필드명 정리
                        clean_name = field_name.replace("<", "").replace(">", "").replace(":", "_").replace(" ", "_")
                        clean_name = clean_name.replace("'", "").replace("(", "").replace(")", "")
                    
                    # 빈 문자열이나 숫자만 있는 경우 비트 위치 기반 이름 생성
                    if not clean_name or clean_name.isdigit():
                        clean_name = f"BIT_{bit_num}"
                
                # 병합된 셀인지 확인
                merge_info = merged_info.get((name_row, bit_col))
                if merge_info:
                    # 병합된 셀의 시작점에서 이름 가져오기
                    master_name = df.iat[merge_info['min_row'], merge_info['min_col']]
                    if pd.notna(master_name):
                        field_name = str(master_name).strip()
                        # 동일한 필드명 정리 로직 적용
                        if field_name.startswith("1'b"):
                            bit_value = field_name.replace("1'b", "")
                            clean_name = f"BIT{bit_value}"
                        else:
                            clean_name = field_name.replace("<", "").replace(">", "").replace(":", "_").replace(" ", "_")
                            clean_name = clean_name.replace("'", "").replace("(", "").replace(")", "")
                        
                        if not clean_name or clean_name.isdigit():
                            # 병합된 필드의 경우 범위 기반 이름 생성
                            bit_range_name = f"{upper_bit}_{lower_bit}" if upper_bit != lower_bit else str(upper_bit)
                            clean_name = f"FIELD_{bit_range_name}"
                    
                    # 이미 처리된 병합 필드인지 확인
                    merge_key = (merge_info['min_row'], merge_info['min_col'], merge_info['max_row'], merge_info['max_col'])
                    if merge_key in processed_merged_fields:
                        continue  # 이미 처리된 병합 필드는 스킵
                    processed_merged_fields.add(merge_key)
                    
                    # 병합 범위 계산 (비트 번호 기준)
                    upper_bit = 15 - (merge_info['min_col'] - addr_col - 1)
                    lower_bit = 15 - (merge_info['max_col'] - addr_col - 1)
                    
                    # upper가 lower보다 작으면 바꿔줌
                    if upper_bit < lower_bit:
                        upper_bit, lower_bit = lower_bit, upper_bit
                        
                else:
                    # 단일 비트
                    upper_bit = lower_bit = bit_num
                
                # Default 값 처리 (병합된 셀도 비트별로 계산)
                default_val = 0
                if merge_info:
                    # 병합된 셀의 경우 각 비트별로 Default 값을 읽어서 계산
                    bit_count = upper_bit - lower_bit + 1
                    calculated_default = 0
                    
                    print(f"      🔍 병합된 필드 '{clean_name}' [{upper_bit}:{lower_bit}] - {bit_count}비트 개별 계산")
                    
                    for bit_pos in range(lower_bit, upper_bit + 1):
                        # 해당 비트 위치의 열 계산
                        bit_col_pos = addr_col + 1 + (15 - bit_pos)
                        if bit_col_pos < len(df.columns) and default_row < len(df):
                            bit_default_cell = df.iat[default_row, bit_col_pos]
                            if pd.notna(bit_default_cell):
                                try:
                                    bit_default_val = int(bit_default_cell)
                                    if bit_default_val != 0:
                                        # 해당 비트 위치에 값 설정
                                        bit_offset = bit_pos - lower_bit
                                        calculated_default |= (bit_default_val << bit_offset)
                                        print(f"        🔸 비트 {bit_pos}: {bit_default_val} -> 오프셋 {bit_offset}")
                                except:
                                    pass
                    
                    default_val = calculated_default
                    print(f"      ✅ 병합된 필드 '{clean_name}' 계산된 Default: {default_val} (0x{default_val:X})")
                else:
                    # 단일 비트의 경우
                    if pd.notna(default_cell):
                        try:
                            default_val = int(default_cell)
                        except:
                            default_val = 0
                    print(f"      🔸 단일 비트 '{clean_name}' Default: {default_val}")
                
                # 필드명이 있는 경우만 추가
                if field_name and field_name not in ["nan", ""] and clean_name:
                    # 같은 이름의 필드가 이미 있는지 확인
                    existing_field = None
                    for field in fields:
                        if field["name"] == clean_name:
                            existing_field = field
                            break
                    
                    if existing_field is None:
                        # 의미 정보 가져오기
                        field_meaning = field_meanings.get(field_name, f"{field_name} bits {upper_bit}:{lower_bit}" if upper_bit != lower_bit else f"{field_name} bit {upper_bit}")
                        
                        # 비트 범위 문자열 생성
                        if upper_bit == lower_bit:
                            bit_range_str = str(upper_bit)
                        else:
                            bit_range_str = f"{upper_bit}:{lower_bit}"
                        
                        field_data = {
                            "name": clean_name,
                            "bit_range": bit_range_str,
                            "upper_bit": upper_bit,
                            "lower_bit": lower_bit,
                            "default_value": str(default_val),
                            "meaning": field_meaning
                        }
                        
                        fields.append(field_data)
                        print(f"     🔹 필드: {clean_name} = bit {upper_bit}:{lower_bit}, 기본값: {default_val}, 의미: {field_meaning}")
            
            if not fields:
                print("   ⚠️ 필드가 발견되지 않음")
                return None
            
            # 레지스터 기본값 계산
            default_value = self.calculate_register_default_value(fields)
            
            register_data = {
                "address": address,
                "description": description,
                "fields": fields,
                "default_value": default_value
            }
            
            print(f"   ✅ 레지스터 파싱 완료: {len(fields)}개 필드, 기본값: {default_value}")
            return register_data
            
        except Exception as e:
            print(f"   ❌ 레지스터 파싱 중 오류: {e}")
            import traceback
            traceback.print_exc()
            return None

    def group_consecutive_fields(self, bit_info):
        """연속된 같은 이름의 필드들을 그룹화합니다."""
        if not bit_info:
            return []
        
        groups = []
        
        # 비트 번호 순으로 정렬 (내림차순 - MSB부터)
        sorted_bits = sorted(bit_info, key=lambda x: x['bit'], reverse=True)
        
        # 이름이 있는 비트와 없는 비트를 구분
        named_bits = [bit for bit in sorted_bits if bit['name'] is not None]
        unnamed_bits = [bit for bit in sorted_bits if bit['name'] is None]
        
        # 이름이 있는 필드들을 처리
        for bit_data in named_bits:
            bit_num = bit_data['bit']
            field_name = str(bit_data['name']).strip()
            
            # 이미 처리된 비트는 건너뛰기
            if any(bit_num in group.get('bits', []) for group in groups):
                continue
            
            # 필드 이름에서 비트 범위 정보 추출 (예: TX_SEN<13:0>, RX0_SEN<15:0>)
            extracted_range = self.extract_bit_range_from_name(field_name)
            
            if extracted_range:
                # 이름에 비트 범위가 명시된 경우
                upper_bit, lower_bit = extracted_range
                clean_name = self.clean_field_name(field_name)
                
                groups.append({
                    'name': clean_name,
                    'max_bit': upper_bit,
                    'min_bit': lower_bit,
                    'default': bit_data['default'],
                    'bits': list(range(upper_bit, lower_bit - 1, -1))
                })
                print(f"      🔍 범위 추출: {field_name} -> {clean_name} = {upper_bit}:{lower_bit}")
                
            else:
                # 일반적인 필드 처리
                same_name_bits = [b for b in named_bits if str(b['name']).strip() == field_name]
                
                if len(same_name_bits) == 1:
                    # 단일 이름의 필드
                    if len(unnamed_bits) >= 15 and bit_num == 15:
                        # 15번 비트에 이름이 있고 나머지가 모두 unnamed이면 전체 필드로 간주 (reset 케이스)
                        groups.append({
                            'name': field_name,
                            'max_bit': 15,
                            'min_bit': 0,
                            'default': bit_data['default'],
                            'bits': list(range(15, -1, -1))
                        })
                    else:
                        # 일반 단일 비트 필드
                        groups.append({
                            'name': field_name,
                            'max_bit': bit_num,
                            'min_bit': bit_num,
                            'default': bit_data['default'],
                            'bits': [bit_num]
                        })
                else:
                    # 같은 이름의 여러 비트들을 연속 그룹으로 처리
                    consecutive_groups = self.find_consecutive_groups(same_name_bits)
                    for group_bits in consecutive_groups:
                        bit_numbers = [b['bit'] for b in group_bits]
                        groups.append({
                            'name': field_name,
                            'max_bit': max(bit_numbers),
                            'min_bit': min(bit_numbers),
                            'default': group_bits[0]['default'],
                            'bits': sorted(bit_numbers, reverse=True)
                        })
        
        return groups
    
    def extract_bit_range_from_name(self, field_name):
        """필드 이름에서 비트 범위를 추출합니다. 예: TX_SEN<13:0> -> (13, 0)"""
        import re
        
        # <숫자:숫자> 패턴 찾기
        pattern = r'<(\d+):(\d+)>'
        match = re.search(pattern, field_name)
        if match:
            upper = int(match.group(1))
            lower = int(match.group(2))
            return (upper, lower)
        
        # <숫자> 패턴 찾기 (단일 비트)
        pattern = r'<(\d+)>'
        match = re.search(pattern, field_name)
        if match:
            bit_num = int(match.group(1))
            return (bit_num, bit_num)
        
        return None
    
    def clean_field_name(self, field_name):
        """필드 이름에서 비트 범위 표기를 제거합니다. 예: TX_SEN<13:0> -> TX_SEN"""
        import re
        # <...> 부분 제거
        cleaned = re.sub(r'<[^>]+>', '', field_name)
        return cleaned.strip()
    
    def find_consecutive_groups(self, bits):
        """같은 이름의 비트들을 연속된 그룹들로 나눕니다."""
        if not bits:
            return []
        
        # 비트 번호로 정렬
        sorted_bits = sorted(bits, key=lambda x: x['bit'], reverse=True)
        
        groups = []
        current_group = [sorted_bits[0]]
        
        for i in range(1, len(sorted_bits)):
            current_bit = sorted_bits[i]['bit']
            prev_bit = sorted_bits[i-1]['bit']
            
            if prev_bit - current_bit == 1:
                # 연속된 비트
                current_group.append(sorted_bits[i])
            else:
                # 연속되지 않음 - 새 그룹 시작
                groups.append(current_group)
                current_group = [sorted_bits[i]]
        
        # 마지막 그룹 추가
        groups.append(current_group)
        
        return groups

    def build_tree(self):
        """트리 구조를 구축합니다."""
        self.ui.tree_widget.clear()
        
        if not self.data:
            return
        
        print(f"🌳 트리 구성 시작, 데이터 구조: {type(self.data)}")
        
        # 데이터 구조에 따라 처리
        if isinstance(self.data, dict):
            for sheet_name, registers in self.data.items():
                print(f"📋 시트: {sheet_name}, 레지스터 수: {len(registers)}")
                
                # 시트 아이템 생성
                sheet_item = QTreeWidgetItem(self.ui.tree_widget, [sheet_name])
                
                for register in registers:
                    # 레지스터 아이템 생성
                    reg_text = f"{register['address']} - {register['description']}"
                    reg_item = QTreeWidgetItem(sheet_item, [reg_text])
                    
                    # 레지스터 아이템을 클릭 가능하게 설정
                    reg_item.setFlags(reg_item.flags() | Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                    
                    # 레지스터 데이터 저장 (주소에서 0x 제거)
                    register_address = register['address'].replace('0x', '') if register['address'].startswith('0x') else register['address']
                    reg_item.setData(0, Qt.UserRole, {
                        'type': 'register',
                        'address': register_address,
                        'description': register['description'],
                        'default_value': register.get('default_value', 0),
                        'fields': register.get('fields', [])
                    })
                    
                    print(f"  📌 레지스터 추가: {register['address']} - {register['description']}")
                    
                    # 필드 아이템들 추가
                    for field in register.get('fields', []):
                        field_text = f"{field['name']} [{field['bit_range']}] = {field['default_value']}"
                        field_item = QTreeWidgetItem(reg_item, [field_text])
                        # 필드 아이템도 클릭 가능하게 설정
                        field_item.setFlags(field_item.flags() | Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                        field_item.setData(0, Qt.UserRole, {
                            'type': 'field',
                            'name': field['name'],
                            'bit_range': field['bit_range'],
                            'default_value': field['default_value'],
                            'meaning': field.get('meaning', '')
                        })
        
        # 트리 확장
        self.ui.tree_widget.expandAll()
        print("✅ 트리 구성 완료")

    def open_excel_file(self):
        """Excel 파일 열기 대화상자"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Excel 파일 선택", "", "Excel Files (*.xlsx *.xls)"
        )
        if file_path:
            self.load_excel_file(file_path)

    def save_json_file(self):
        """JSON 파일 저장 대화상자"""
        if not self.data:
            QMessageBox.warning(self, "경고", "저장할 데이터가 없습니다.")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "JSON 파일 저장", "", "JSON Files (*.json)"
        )
        if file_path:
            try:
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(self.data, f, ensure_ascii=False, indent=2)
                QMessageBox.information(self, "성공", f"파일이 저장되었습니다:\\n{file_path}")
            except Exception as e:
                QMessageBox.critical(self, "저장 오류", f"파일 저장 실패:\\n{str(e)}")

    def update_tree_display_values(self, new_value):
        """현재 선택된 레지스터의 Tree 표시 값만 업데이트합니다."""
        try:
            if not self.current_register or not hasattr(self, 'ui') or not hasattr(self.ui, 'tree_widget'):
                return
                
            print(f"🌳 Tree 값 업데이트 시작 (레지스터 {self.current_register}): 0x{new_value:08X}")
            
            # 현재 선택된 레지스터만 찾기
            root = self.ui.tree_widget.invisibleRootItem()
            target_register_item = None
            
            for sheet_idx in range(root.childCount()):
                sheet_item = root.child(sheet_idx)
                for reg_idx in range(sheet_item.childCount()):
                    reg_item = sheet_item.child(reg_idx)
                    reg_data = reg_item.data(0, Qt.UserRole)
                    
                    if reg_data and reg_data.get('address') == self.current_register:
                        target_register_item = reg_item
                        break
                if target_register_item:
                    break
            
            if not target_register_item:
                print(f"⚠️ 현재 레지스터 {self.current_register}를 Tree에서 찾을 수 없음")
                return
            
            # 현재 선택된 레지스터의 데이터 가져오기
            reg_data = target_register_item.data(0, Qt.UserRole)
            
            # 레지스터 아이템 텍스트 업데이트 (값 표시 없이)
            reg_text = f"0x{self.current_register} - {reg_data.get('description', '')}"
            target_register_item.setText(0, reg_text)
            print(f"📋 레지스터 아이템 업데이트: {reg_text} (값: {new_value})")
            
            # 현재 선택된 레지스터의 각 필드 값 계산 및 업데이트
            for field_idx in range(target_register_item.childCount()):
                field_item = target_register_item.child(field_idx)
                field_data = field_item.data(0, Qt.UserRole)
                
                if field_data and field_data.get('type') == 'field':
                    field_name = field_data.get('name', '')
                    bit_range = field_data.get('bit_range', '')
                    
                    # 비트 범위 파싱 (공용 함수 사용)
                    upper_bit, lower_bit = self.parse_bit_range(bit_range)
                    
                    # 해당 비트 범위의 값 추출 (공용 함수 사용)
                    field_value = self.extract_field_value_from_register(new_value, upper_bit, lower_bit)
                    
                    # 현재 선택된 필드인지 확인하여 추가 디버깅
                    if self.current_field_data and field_name == self.current_field_data.get('name'):
                        print(f"  🎯 현재 선택된 필드 업데이트: {field_name}")
                        print(f"     비트 범위: {upper_bit}:{lower_bit}")
                        print(f"     레지스터 값: 0x{new_value:08X}")
                        print(f"     계산된 필드 값: {field_value}")
                    
                    # 필드 아이템 텍스트 업데이트
                    field_text = f"{field_name} [{bit_range}] = {field_value}"
                    field_item.setText(0, field_text)
                    print(f"  🔹 필드 업데이트: {field_text}")
            
            print(f"✅ Tree 값 업데이트 완료 (레지스터 {self.current_register}만)")
            
        except Exception as e:
            print(f"❌ Tree 값 업데이트 중 오류: {e}")
            import traceback
            traceback.print_exc()

    def show_about(self):
        """정보 대화상자"""
        QMessageBox.about(self, "정보", 
                         "Register Tree Viewer & Controller\\n\\n"
                         "Excel 파일에서 레지스터 정보를 읽어와 트리 구조로 표시하고\\n"
                         "FT2232H를 통한 SPI 통신으로 레지스터를 제어합니다.")

def main():
    app = QApplication(sys.argv)
    
    # Excel 파일 경로 (있으면 자동 로드)
    excel_path = "Sample.xlsx"
    
    try:
        window = RegisterTreeViewerController(excel_path)
        window.show()
        sys.exit(app.exec())
    except Exception as e:
        print(f"애플리케이션 시작 오류: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
